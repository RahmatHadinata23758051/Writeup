from collections import defaultdict
import math

from openpyxl import load_workbook


WORKBOOK = "challenge.xlsx"


def extract_flag(path: str) -> str:
    wb = load_workbook(path, data_only=False)
    ws = wb["Network"]

    constraints = defaultdict(lambda: {"ge": [], "le": []})

    for idx, row in enumerate(range(11, 71)):
        pos = None
        weight = None
        for col in range(3, 33):
            value = ws.cell(row, col).value
            if isinstance(value, (int, float)) and abs(value) > 1e-12:
                pos = col - 2
                weight = value
                break

        if pos is None:
            continue

        bias = ws.cell(75, idx + 3).value
        ascii_threshold = round(abs(bias) * 127)

        if weight == 1:
            constraints[pos]["le"].append(ascii_threshold)
        elif weight == -1:
            constraints[pos]["ge"].append(ascii_threshold)

    chars = []
    for pos in range(1, 31):
        lower = max(constraints[pos]["ge"]) if constraints[pos]["ge"] else 0
        upper = min(constraints[pos]["le"]) if constraints[pos]["le"] else 0
        if lower != upper:
            raise ValueError(f"ambiguous constraint at position {pos}: {lower=} {upper=}")
        if lower == 0:
            break
        chars.append(chr(lower))

    return "".join(chars)


def validate_flag(path: str, flag: str) -> bool:
    wb = load_workbook(path, data_only=False)
    ws = wb["Network"]

    x = [ord(c) / 127 for c in flag] + [0.0] * (30 - len(flag))

    a1 = []
    for idx, row in enumerate(range(11, 71)):
        total = 0.0
        for col in range(3, 33):
            weight = ws.cell(row, col).value or 0
            total += x[col - 3] * weight
        total += ws.cell(75, idx + 3).value
        a1.append(max(0.0, total))

    z2 = sum(a1[i] * ws.cell(87, i + 3).value for i in range(60)) + ws["C92"].value
    a2 = max(0.0, z2)

    z3 = [
        a2 * ws["C101"].value + ws["C104"].value,
        a2 * ws["D101"].value + ws["D104"].value,
        a2 * ws["E101"].value + ws["E104"].value,
        a2 * ws["F101"].value + ws["F104"].value,
    ]
    a3 = [1 / (1 + math.exp(-value)) for value in z3]

    return a3[0] > 0.5 and a3[1] < 0.5 and a3[2] > 0.5 and a3[3] < 0.5


def main() -> None:
    flag = extract_flag(WORKBOOK)
    if not validate_flag(WORKBOOK, flag):
        raise SystemExit("validation failed")
    print(flag)


if __name__ == "__main__":
    main()
